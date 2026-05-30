from __future__ import annotations

import csv
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
from pathlib import Path
import subprocess
import sys
import tempfile
from threading import Thread
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"
SITE_INDEX = FIXTURES / "demo_site" / "index.html"


class TradeAutomationTests(unittest.TestCase):
    def run_script(self, script: str, *args: str) -> subprocess.CompletedProcess[str]:
        command = [sys.executable, str(ROOT / "tools" / script), *args]
        return subprocess.run(command, cwd=ROOT, capture_output=True, text=True)

    def test_collect_prospects_generates_search_tasks_without_api(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            result = self.run_script(
                "collect_prospects.py",
                "--discovery",
                str(ROOT / "templates" / "DISCOVERY.example.yaml"),
                "--product",
                str(ROOT / "templates" / "PRODUCT.example.yaml"),
                "--output-dir",
                str(output_dir),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            tasks_csv = output_dir / "prospect_search_tasks.csv"
            tasks_json = output_dir / "prospect_search_tasks.json"
            self.assertTrue(tasks_csv.exists())
            self.assertTrue(tasks_json.exists())

            with tasks_csv.open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertGreaterEqual(len(rows), 1)
            self.assertIn("folding camping table", {row["keyword"].lower() for row in rows})
            self.assertIn("United States", {row["region"] for row in rows})

    def test_collect_prospects_calls_configured_api(self) -> None:
        class Handler(BaseHTTPRequestHandler):
            def do_GET(self) -> None:  # noqa: N802 - stdlib callback name.
                body = {
                    "items": [
                        {
                            "company_name": "API Outdoor Buyer",
                            "website": "https://api-buyer.example",
                            "country": "United States",
                            "business_type": "distributor",
                            "source_url": "http://local-api/source",
                        }
                    ]
                }
                payload = json.dumps(body).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)

            def log_message(self, format: str, *args: object) -> None:
                return

        server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        thread = Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir)
                discovery = temp_path / "DISCOVERY.yaml"
                discovery.write_text(
                    f"""
collection_api:
  provider: "local"
  endpoint: "http://127.0.0.1:{server.server_port}/search"
  method: "GET"
  query_params:
    query: "{{keyword}}"
    region: "{{region}}"
  response_mapping:
    items_path: "items"
    company_name: "company_name"
    website: "website"
    country: "country"
    business_type: "business_type"
    source_url: "source_url"
  pagination:
    max_pages: 1
    page_size: 10
  rate_limit:
    requests_per_minute: 0
default_regions:
  - "United States"
""",
                    encoding="utf-8",
                )
                output_dir = temp_path / "api"
                result = self.run_script(
                    "collect_prospects.py",
                    "--discovery",
                    str(discovery),
                    "--product",
                    str(ROOT / "templates" / "PRODUCT.example.yaml"),
                    "--output-dir",
                    str(output_dir),
                )

                self.assertEqual(result.returncode, 0, result.stderr)
                with (output_dir / "prospects.raw.csv").open(encoding="utf-8") as handle:
                    rows = list(csv.DictReader(handle))
                self.assertEqual(rows[0]["company_name"], "API Outdoor Buyer")
        finally:
            server.shutdown()
            server.server_close()

    def test_decision_maker_finder_extracts_roles_and_email_status(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "decision_makers.json"
            result = self.run_script(
                "decision_maker_finder.py",
                "--website",
                SITE_INDEX.as_uri(),
                "--output",
                str(output_path),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            data = json.loads(output_path.read_text(encoding="utf-8"))
            roles = {item["role"] for item in data["candidates"]}
            self.assertIn("Purchasing Manager", roles)
            email_items = [item for item in data["candidates"] if item.get("email")]
            self.assertEqual(email_items[0]["email_status"], "domain_match")

    def test_decision_maker_finder_uses_configured_contact_api(self) -> None:
        class Handler(BaseHTTPRequestHandler):
            def do_GET(self) -> None:  # noqa: N802 - stdlib callback name.
                body = {
                    "items": [
                        {
                            "name": "Sarah Buyer",
                            "role": "Sourcing Manager",
                            "email": "sarah@demooutdoor.test",
                            "source_url": "http://local-api/contact",
                        }
                    ]
                }
                payload = json.dumps(body).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)

            def log_message(self, format: str, *args: object) -> None:
                return

        server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        thread = Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir)
                discovery = temp_path / "DISCOVERY.yaml"
                discovery.write_text(
                    f"""
contact_enrichment_api:
  provider: "local"
  endpoint: "http://127.0.0.1:{server.server_port}/contacts"
  method: "GET"
  query_params:
    domain: "{{domain}}"
    company_name: "{{company_name}}"
  response_mapping:
    items_path: "items"
    name: "name"
    role: "role"
    email: "email"
    source_url: "source_url"
  rate_limit:
    requests_per_minute: 0
""",
                    encoding="utf-8",
                )
                output_path = temp_path / "decision_makers.json"
                result = self.run_script(
                    "decision_maker_finder.py",
                    "--website",
                    SITE_INDEX.as_uri(),
                    "--company-name",
                    "Demo Outdoor",
                    "--discovery",
                    str(discovery),
                    "--output",
                    str(output_path),
                )

                self.assertEqual(result.returncode, 0, result.stderr)
                data = json.loads(output_path.read_text(encoding="utf-8"))
                api_candidates = [item for item in data["candidates"] if item["email_status"] == "api_verified"]
                self.assertEqual(api_candidates[0]["name"], "Sarah Buyer")
        finally:
            server.shutdown()
            server.server_close()

    def test_decision_maker_finder_forces_contact_search_and_reports_none(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            contact_site = temp_path / "contact-only.html"
            contact_site.write_text(
                "<!doctype html><html><body>"
                "<h1>Contact Export Supply</h1>"
                "<p>Email: sales@exportsupply.test</p>"
                "<p>Phone: +1 555-010-9000</p>"
                "</body></html>",
                encoding="utf-8",
            )
            output_path = temp_path / "decision_makers.json"

            result = self.run_script(
                "decision_maker_finder.py",
                "--website",
                contact_site.as_uri(),
                "--output",
                str(output_path),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            data = json.loads(output_path.read_text(encoding="utf-8"))
            self.assertEqual(data["contact_search"]["email_result"], "found")
            self.assertEqual(data["contact_search"]["phone_result"], "found")
            self.assertEqual(data["contact_search"]["emails"][0]["value"], "sales@exportsupply.test")
            self.assertEqual(data["contact_search"]["phones"][0]["value"], "+1 555-010-9000")

            no_contact_site = temp_path / "no-contact.html"
            no_contact_site.write_text(
                "<!doctype html><html><body><h1>No Contact Listed</h1></body></html>",
                encoding="utf-8",
            )
            no_contact_output = temp_path / "no_contact_decision_makers.json"
            result = self.run_script(
                "decision_maker_finder.py",
                "--website",
                no_contact_site.as_uri(),
                "--output",
                str(no_contact_output),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            no_contact_data = json.loads(no_contact_output.read_text(encoding="utf-8"))
            self.assertEqual(no_contact_data["contact_search"]["email_result"], "没有")
            self.assertEqual(no_contact_data["contact_search"]["phone_result"], "没有")
            self.assertEqual(no_contact_data["contact_search"]["emails"], [])
            self.assertEqual(no_contact_data["contact_search"]["phones"], [])

    def test_batch_pipeline_outputs_enriched_workbooks_and_email_drafts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_csv = temp_path / "prospects.csv"
            input_csv.write_text(
                "company_name,website,country,source_url\n"
                f"Demo Outdoor,{SITE_INDEX.as_uri()},United States,fixture-a\n"
                f"Demo Outdoor LLC,{SITE_INDEX.as_uri()},United States,fixture-b\n"
                "Missing Website,,United States,fixture-c\n",
                encoding="utf-8",
            )
            output_dir = temp_path / "pipeline"

            result = self.run_script(
                "batch_prospect_pipeline.py",
                "--input",
                str(input_csv),
                "--product",
                str(ROOT / "templates" / "PRODUCT.example.yaml"),
                "--market",
                str(ROOT / "templates" / "MARKET.example.yaml"),
                "--tone",
                str(ROOT / "templates" / "TONE.example.yaml"),
                "--output-dir",
                str(output_dir),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            enriched = output_dir / "prospects.enriched.xlsx"
            scores = output_dir / "scores.xlsx"
            emails = output_dir / "email_drafts.xlsx"
            reports = output_dir / "research_reports.json"
            for path in [enriched, scores, emails, reports]:
                self.assertTrue(path.exists(), f"missing {path}")

            score_book = load_workbook(scores)
            self.assertEqual(score_book["Scores"]["B2"].value, "A")
            email_book = load_workbook(emails)
            body = email_book["Email Drafts"]["D2"].value
            self.assertIn("camping furniture", body)
            self.assertIn("Folding Camping Table", body)

    def test_batch_pipeline_blocks_email_when_product_evidence_is_missing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            unrelated_site = temp_path / "unrelated.html"
            unrelated_site.write_text(
                "<!doctype html><html><body>"
                "<h1>Metro Accounting Group</h1>"
                "<p>We provide tax filing, bookkeeping, and payroll services for local companies.</p>"
                "</body></html>",
                encoding="utf-8",
            )
            discovery = temp_path / "DISCOVERY.yaml"
            discovery.write_text("scraping:\n  engine: \"http\"\n", encoding="utf-8")
            input_csv = temp_path / "prospects.csv"
            input_csv.write_text(
                "company_name,website,country,source_url\n"
                f"Metro Accounting,{unrelated_site.as_uri()},United States,fixture-unrelated\n",
                encoding="utf-8",
            )
            output_dir = temp_path / "pipeline"

            result = self.run_script(
                "batch_prospect_pipeline.py",
                "--input",
                str(input_csv),
                "--product",
                str(ROOT / "templates" / "PRODUCT.example.yaml"),
                "--market",
                str(ROOT / "templates" / "MARKET.example.yaml"),
                "--tone",
                str(ROOT / "templates" / "TONE.example.yaml"),
                "--discovery",
                str(discovery),
                "--output-dir",
                str(output_dir),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            email_book = load_workbook(output_dir / "email_drafts.xlsx")
            email_sheet = email_book["Email Drafts"]
            email_headers = [cell.value for cell in email_sheet[1]]
            email_row = dict(zip(email_headers, [cell.value for cell in email_sheet[2]]))
            self.assertEqual(email_row["draft_status"], "blocked_no_evidence")
            self.assertNotIn("works in a related outdoor category", email_row["body"])
            self.assertIn("No product overlap was verified", email_row["body"])

            score_book = load_workbook(output_dir / "scores.xlsx")
            score_headers = [cell.value for cell in score_book["Scores"][1]]
            score_row = dict(zip(score_headers, [cell.value for cell in score_book["Scores"][2]]))
            self.assertEqual(score_row["recommended_action"], "manual_review")

            reports = json.loads((output_dir / "research_reports.json").read_text(encoding="utf-8"))
            self.assertEqual(reports[0]["evidence_status"], "no_product_evidence")


if __name__ == "__main__":
    unittest.main()
