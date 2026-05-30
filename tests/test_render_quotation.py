from __future__ import annotations

import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "tools" / "render_quotation.py"
EXAMPLE = ROOT / "examples" / "quotation.example.json"


class RenderQuotationTests(unittest.TestCase):
    def run_renderer(self, input_path: Path, output_dir: Path, *formats: str) -> subprocess.CompletedProcess[str]:
        command = [
            sys.executable,
            str(SCRIPT),
            str(input_path),
            "--output-dir",
            str(output_dir),
            "--formats",
            *formats,
        ]
        return subprocess.run(command, cwd=ROOT, capture_output=True, text=True)

    def test_renders_html_and_excel_exports(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            result = self.run_renderer(EXAMPLE, output_dir, "html", "excel")

            self.assertEqual(result.returncode, 0, result.stderr)
            html_path = output_dir / "quotation-Q-2026-001-demo-outdoor-inc-draft.html"
            excel_path = output_dir / "quotation-Q-2026-001-demo-outdoor-inc-draft.xlsx"
            self.assertTrue(html_path.exists())
            self.assertTrue(excel_path.exists())

            html = html_path.read_text(encoding="utf-8")
            self.assertIn("Q-2026-001", html)
            self.assertIn("Demo Outdoor Inc.", html)
            self.assertIn("CT-200A", html)
            self.assertIn("8,600.00", html)

            workbook = load_workbook(excel_path)
            self.assertEqual(workbook.sheetnames, ["Quotation", "Items", "Terms", "Review Notes"])
            self.assertEqual(workbook["Items"]["A2"].value, "CT-200A")
            self.assertEqual(workbook["Quotation"]["B6"].value, 8600)

    def test_blocked_quotation_does_not_export(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            blocked_input = temp_path / "blocked.json"
            blocked_input.write_text(
                json.dumps(
                    {
                        "quotation_status": "blocked",
                        "quotation_number": "Q-BLOCKED",
                        "buyer": {"name": "Blocked Buyer"},
                        "missing_fields": ["unit_price"],
                    }
                ),
                encoding="utf-8",
            )

            output_dir = temp_path / "exports"
            result = self.run_renderer(blocked_input, output_dir, "html", "excel")

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("blocked", result.stderr.lower())
            self.assertFalse(output_dir.exists())


if __name__ == "__main__":
    unittest.main()
