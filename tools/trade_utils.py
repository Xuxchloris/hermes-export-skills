from __future__ import annotations

import csv
import json
import os
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
import re
import time
from typing import Any
from urllib.parse import urljoin, urlparse
from urllib.request import ProxyHandler, Request, build_opener, urlopen

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


class LinkAndTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[str] = []
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "a":
            for name, value in attrs:
                if name.lower() == "href" and value:
                    self.links.append(value)

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(data.strip())

    @property
    def text(self) -> str:
        return "\n".join(self.parts)


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def load_json_path(path: str, data: Any) -> Any:
    current = data
    for part in path.split("."):
        if part == "":
            continue
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, list) and part.isdigit():
            current = current[int(part)]
        else:
            return None
    return current


def render_template(value: Any, variables: dict[str, Any]) -> Any:
    if isinstance(value, str):
        rendered = value
        for key, replacement in variables.items():
            rendered = rendered.replace("{" + key + "}", str(replacement))
        return rendered
    if isinstance(value, dict):
        return {key: render_template(item, variables) for key, item in value.items()}
    if isinstance(value, list):
        return [render_template(item, variables) for item in value]
    return value


def _product_tokens(product: dict[str, Any]) -> list[str]:
    tokens: list[str] = []
    for key in ["sku", "name", "category", "hs_code"]:
        value = product.get(key)
        if value:
            tokens.append(str(value))
    for key in ["keywords", "target_applications"]:
        tokens.extend(str(item) for item in product.get(key, []))
    return [token.lower() for token in tokens if token]


def select_product_config(product_config: dict[str, Any], product_query: str = "", sku: str = "") -> dict[str, Any]:
    products = list(product_config.get("products", []))
    if not products:
        return product_config

    query = product_query.strip().lower()
    sku = sku.strip().lower()
    if not query and not sku:
        return product_config

    matched: list[dict[str, Any]] = []
    for product in products:
        tokens = _product_tokens(product)
        if sku and str(product.get("sku", "")).lower() == sku:
            matched = [product]
            break
        if query and any(query in token or token in query for token in tokens):
            matched.append(product)

    if not matched:
        raise ValueError(f"No product matched query '{product_query or sku}'")

    selected = dict(product_config)
    selected["products"] = matched
    selected["selected_product_query"] = product_query
    selected["selected_product_sku"] = sku
    return selected


def scraping_proxy_options(scraping: dict[str, Any] | None = None) -> dict[str, Any]:
    scraping = scraping or {}
    proxy = str(scraping.get("proxy") or os.environ.get("SCRAPING_PROXY_URL") or "").strip()
    proxies = dict(scraping.get("proxies") or {})
    http_proxy = str(os.environ.get("SCRAPING_PROXY_HTTP") or "").strip()
    https_proxy = str(os.environ.get("SCRAPING_PROXY_HTTPS") or "").strip()

    if not proxies and (http_proxy or https_proxy):
        proxies = {}
        if http_proxy:
            proxies["http"] = http_proxy
        if https_proxy:
            proxies["https"] = https_proxy

    if proxy:
        return {"proxy": proxy}
    if proxies:
        return {"proxies": proxies}
    return {}


def scrapling_fetch(url: str, scraping: dict[str, Any]) -> str:
    engine = scraping.get("engine", "scrapling-fetcher")
    proxy_options = scraping_proxy_options(scraping)
    try:
        if engine == "scrapling-fetcher":
            from scrapling.fetchers import Fetcher

            page = Fetcher.get(url, stealthy_headers=bool(scraping.get("stealthy_headers", True)), **proxy_options)
        elif engine == "scrapling-dynamic":
            from scrapling.fetchers import DynamicFetcher

            page = DynamicFetcher.fetch(
                url,
                headless=bool(scraping.get("headless", True)),
                network_idle=bool(scraping.get("network_idle", True)),
                **proxy_options,
            )
        elif engine == "scrapling-stealthy":
            from scrapling.fetchers import StealthyFetcher

            page = StealthyFetcher.fetch(
                url,
                headless=bool(scraping.get("headless", True)),
                network_idle=bool(scraping.get("network_idle", True)),
                **proxy_options,
            )
        else:
            raise ValueError(f"Unknown scraping engine: {engine}")
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Scrapling is the default scraping backend. Install dependencies with "
            "`python -m pip install -r requirements.txt`; "
            "run `scrapling install` for browser-based engines."
        ) from error
    return str(getattr(page, "html_content", page))


def fetch_url(url: str, scraping: dict[str, Any] | None = None, timeout: int = 12) -> str:
    scraping = scraping or {}
    engine = scraping.get("engine", "scrapling-fetcher")
    if engine.startswith("scrapling-"):
        return scrapling_fetch(url, scraping)

    request = Request(url, headers={"User-Agent": "HermesTradeAgent/0.1"})
    proxy_options = scraping_proxy_options(scraping)
    if "proxy" in proxy_options:
        proxy_url = proxy_options["proxy"]
        opener = build_opener(ProxyHandler({"http": proxy_url, "https": proxy_url}))
        response_context = opener.open(request, timeout=timeout)
    elif "proxies" in proxy_options:
        response_context = build_opener(ProxyHandler(proxy_options["proxies"])).open(request, timeout=timeout)
    else:
        response_context = urlopen(request, timeout=timeout)
    with response_context as response:
        content_type = response.headers.get("content-type", "")
        charset = "utf-8"
        match = re.search(r"charset=([\w-]+)", content_type)
        if match:
            charset = match.group(1)
        return response.read().decode(charset, errors="replace")


def parse_html(html: str) -> tuple[str, list[str]]:
    parser = LinkAndTextParser()
    parser.feed(html)
    text = unescape(re.sub(r"\n{3,}", "\n\n", parser.text))
    return text, parser.links


def crawl_company_pages(website: str, max_pages: int = 4, scraping: dict[str, Any] | None = None) -> list[dict[str, str]]:
    queue = [website]
    seen: set[str] = set()
    pages: list[dict[str, str]] = []
    preferred = ("about", "contact", "team", "company", "catalog", "product")

    while queue and len(pages) < max_pages:
        url = queue.pop(0)
        if url in seen:
            continue
        seen.add(url)
        try:
            html = fetch_url(url, scraping)
        except Exception as error:  # noqa: BLE001 - record fetch failure for review output.
            pages.append({"url": url, "text": "", "error": str(error)})
            continue

        text, links = parse_html(html)
        pages.append({"url": url, "text": text, "error": ""})
        for link in links:
            absolute = urljoin(url, link)
            parsed = urlparse(absolute)
            if parsed.scheme not in {"http", "https", "file"}:
                continue
            if absolute in seen or absolute in queue:
                continue
            if any(token in absolute.lower() for token in preferred):
                queue.append(absolute)
    return pages


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def read_table(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path)
        sheet = workbook.active
        headers = [normalize_header(str(cell.value or "")) for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            if any(value not in (None, "") for value in record.values()):
                rows.append(record)
        return rows

    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return [{normalize_header(key): value for key, value in row.items()} for row in reader]


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first = True
    for title, rows in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet(title)
        sheet.title = title
        first = False
        headers = list(rows[0].keys()) if rows else ["status"]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def website_key(value: str) -> str:
    if not value:
        return ""
    parsed = urlparse(value if "://" in value else "https://" + value)
    if parsed.scheme == "file":
        return parsed.path.lower()
    host = parsed.netloc.lower().removeprefix("www.")
    return host or value.lower().strip()


def sleep_for_rate_limit(requests_per_minute: int) -> None:
    if requests_per_minute > 0:
        time.sleep(min(60 / requests_per_minute, 1.0))
