from __future__ import annotations

import argparse
from html import escape
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "skills" / "quotation-generator" / "templates" / "quotation.html"


def text(value: Any) -> str:
    return "" if value is None else str(value)


def money(value: Any) -> str:
    return f"{float(value or 0):,.2f}"


def slug(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return cleaned or "buyer"


def validate_quotation(quotation: dict[str, Any]) -> None:
    if quotation.get("quotation_status") == "blocked":
        missing = ", ".join(quotation.get("missing_fields", [])) or "review required"
        raise ValueError(f"Quotation is blocked: {missing}")

    required = ["quotation_number", "buyer", "items", "terms", "total_amount"]
    missing = [field for field in required if field not in quotation]
    if missing:
        raise ValueError(f"Quotation is missing required fields: {', '.join(missing)}")

    if not quotation["items"]:
        raise ValueError("Quotation must contain at least one item")


def output_stem(quotation: dict[str, Any]) -> str:
    buyer = quotation.get("buyer", {})
    return f"quotation-{quotation['quotation_number']}-{slug(text(buyer.get('name')))}-draft"


def item_rows(items: list[dict[str, Any]]) -> str:
    rows = []
    for item in items:
        values = [
            item.get("sku"),
            item.get("product"),
            item.get("specification"),
            item.get("packing"),
            item.get("quantity"),
            money(item.get("unit_price")),
            money(item.get("amount")),
        ]
        cells = "".join(f"<td>{escape(text(value))}</td>" for value in values)
        rows.append(f"      <tr>{cells}</tr>")
    return "\n".join(rows)


def render_html(quotation: dict[str, Any], template_path: Path, output_path: Path) -> None:
    buyer = quotation.get("buyer", {})
    seller = quotation.get("seller", {})
    terms = quotation.get("terms", {})
    review_notes = quotation.get("review_notes", [])
    replacements = {
        "quotation_number": quotation.get("quotation_number"),
        "buyer_name": buyer.get("name"),
        "buyer_country": buyer.get("country"),
        "seller_name": seller.get("name"),
        "quotation_date": quotation.get("quotation_date"),
        "validity": terms.get("validity"),
        "incoterm": terms.get("incoterm"),
        "payment_terms": terms.get("payment_terms"),
        "lead_time": terms.get("lead_time"),
        "items": item_rows(quotation.get("items", [])),
        "total_amount": f"{escape(text(terms.get('currency')))} {money(quotation.get('total_amount'))}".strip(),
        "review_notes": "<br>".join(escape(text(note)) for note in review_notes) or "None",
        "human_review_required": str(bool(quotation.get("human_review_required", True))).lower(),
    }

    html = template_path.read_text(encoding="utf-8")
    for key, value in replacements.items():
        rendered = text(value) if key in {"items", "review_notes", "total_amount"} else escape(text(value))
        html = html.replace("{{" + key + "}}", rendered)
    output_path.write_text(html, encoding="utf-8")


def render_excel(quotation: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Quotation"

    buyer = quotation.get("buyer", {})
    seller = quotation.get("seller", {})
    terms = quotation.get("terms", {})
    summary_rows = [
        ("Quotation Number", quotation.get("quotation_number")),
        ("Buyer", buyer.get("name")),
        ("Buyer Country", buyer.get("country")),
        ("Seller", seller.get("name")),
        ("Currency", terms.get("currency")),
        ("Total Amount", quotation.get("total_amount")),
        ("Status", quotation.get("quotation_status")),
        ("Human Review Required", quotation.get("human_review_required", True)),
    ]
    for row in summary_rows:
        summary.append(row)

    items_sheet = workbook.create_sheet("Items")
    items_sheet.append(["SKU", "Product", "Specification", "Packing", "Quantity", "Unit Price", "Amount"])
    for item in quotation.get("items", []):
        items_sheet.append(
            [
                item.get("sku"),
                item.get("product"),
                item.get("specification"),
                item.get("packing"),
                item.get("quantity"),
                item.get("unit_price"),
                item.get("amount"),
            ]
        )

    terms_sheet = workbook.create_sheet("Terms")
    for key, value in terms.items():
        terms_sheet.append([key, value])

    review_sheet = workbook.create_sheet("Review Notes")
    review_sheet.append(["Type", "Note"])
    for field in quotation.get("missing_fields", []):
        review_sheet.append(["Missing Field", field])
    for note in quotation.get("review_notes", []):
        review_sheet.append(["Review Note", note])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
        for column_cells in sheet.columns:
            width = min(max(len(text(cell.value)) for cell in column_cells) + 2, 55)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)


def browser_candidates() -> list[str]:
    candidates = [
        shutil.which("msedge"),
        shutil.which("chrome"),
        shutil.which("google-chrome"),
        shutil.which("chromium"),
        shutil.which("chromium-browser"),
    ]
    for base in [os.environ.get("PROGRAMFILES"), os.environ.get("PROGRAMFILES(X86)")]:
        if base:
            candidates.extend(
                [
                    str(Path(base) / "Microsoft" / "Edge" / "Application" / "msedge.exe"),
                    str(Path(base) / "Google" / "Chrome" / "Application" / "chrome.exe"),
                ]
            )
    return [candidate for candidate in candidates if candidate and Path(candidate).exists()]


def render_pdf(html_path: Path, output_path: Path) -> None:
    try:
        from weasyprint import HTML

        HTML(filename=str(html_path)).write_pdf(str(output_path))
        return
    except ModuleNotFoundError:
        pass

    browsers = browser_candidates()
    if not browsers:
        raise RuntimeError("PDF export requires WeasyPrint or a local Chrome/Edge browser")

    failures: list[str] = []
    for browser in browsers:
        with tempfile.TemporaryDirectory() as profile_dir:
            command = [
                browser,
                "--headless=new",
                "--disable-gpu",
                "--disable-crash-reporter",
                "--disable-breakpad",
                "--no-first-run",
                "--no-default-browser-check",
                f"--user-data-dir={profile_dir}",
                f"--print-to-pdf={output_path.resolve()}",
                html_path.resolve().as_uri(),
            ]
            result = subprocess.run(command, capture_output=True)
            if output_path.exists():
                return
            stderr = result.stderr.decode("utf-8", errors="replace").strip()
            failures.append(f"{Path(browser).name}: exit {result.returncode}; {stderr[:240]}")

    raise RuntimeError("PDF converter finished without creating a PDF file. " + " | ".join(failures))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render an approved quotation draft to HTML, Excel, and PDF")
    parser.add_argument("quotation_json", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("exports"))
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument(
        "--formats",
        nargs="+",
        choices=["html", "excel", "pdf"],
        default=["html", "excel"],
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        quotation = json.loads(args.quotation_json.read_text(encoding="utf-8"))
        validate_quotation(quotation)
        args.output_dir.mkdir(parents=True, exist_ok=True)
        stem = output_stem(quotation)
        html_path = args.output_dir / f"{stem}.html"

        if "html" in args.formats or "pdf" in args.formats:
            render_html(quotation, args.template, html_path)
            print(f"HTML: {html_path}")
        if "excel" in args.formats:
            excel_path = args.output_dir / f"{stem}.xlsx"
            render_excel(quotation, excel_path)
            print(f"Excel: {excel_path}")
        if "pdf" in args.formats:
            pdf_path = args.output_dir / f"{stem}.pdf"
            render_pdf(html_path, pdf_path)
            print(f"PDF: {pdf_path}")
    except (OSError, ValueError, RuntimeError, json.JSONDecodeError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
